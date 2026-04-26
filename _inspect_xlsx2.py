import openpyxl

wb = openpyxl.load_workbook(
    r'Z:\Prj\ThiagoPaulino\FlávioAnalista\docs\gas_vapor\Tabela_de_classificação_Rev Final.xlsx',
    data_only=True
)

ws = wb['Planilha1']

# Print all unique values per column to understand the data domain
from collections import defaultdict
col_values = defaultdict(set)

for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=False):
    for cell in row:
        v = cell.value
        if v is not None:
            col_letter = cell.column_letter
            val_str = str(v).strip()
            if len(val_str) < 80:
                col_values[col_letter].add(val_str)

print("=== UNIQUE VALUES PER COLUMN (data rows 6-149) ===\n")
col_headers = {
    'A': 'Identificação',
    'B': 'Descrição',
    'C': 'Locação',
    'D': 'Substância Combustível',
    'E': '(merged with D)',
    'F': 'Temp (°C)',
    'G': 'Pressão (kPa)',
    'H': 'Volume (m³)',
    'I': 'Ventilação Tipo',
    'J': 'Ventilação Grau',
    'K': 'Ventilação Disponibilidade',
    'L': 'Fonte Liberação Descrição',
    'M': 'Fonte Liberação Grau',
    'N': 'Grupo-Classe Temp',
    'O': 'Zona 0',
    'P': 'Zona 1 (m)',
    'Q': 'Zona 2 (m)',
    'R': 'Zona 2 adicional (m)',
    'S': 'Zona 20',
    'T': 'Zona 21 (m)',
    'U': 'Zona 22 (m)',
}

for col in 'ABCDEFGHIJKLMNOPQRSTU':
    header = col_headers.get(col, '???')
    vals = sorted(col_values.get(col, set()))
    if col in ('A', 'B'):
        print(f"Col {col} ({header}): {len(vals)} unique values (too many to list)")
    else:
        print(f"Col {col} ({header}): {vals}")
    print()

# Print rows 100-149 to see different area types
print("\n=== ROWS 100-149 (tail of data) ===\n")
for row in ws.iter_rows(min_row=100, max_row=min(ws.max_row, 149), values_only=False):
    vals = []
    for cell in row:
        v = cell.value
        if v is not None:
            vals.append(f'[{cell.coordinate}]={repr(v)}')
    if vals:
        sep = ' | '
        print(f'Row {row[0].row}: {sep.join(vals)}')

# Count non-empty rows
data_rows = 0
for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
    if any(v is not None for v in row):
        data_rows += 1
print(f"\nTotal data rows: {data_rows}")
