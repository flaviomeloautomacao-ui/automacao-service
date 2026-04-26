import openpyxl
import sys

wb = openpyxl.load_workbook(
    r'Z:\Prj\ThiagoPaulino\FlávioAnalista\docs\gas_vapor\Tabela_de_classificação_Rev Final.xlsx',
    data_only=True
)
print('Sheets:', wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== Sheet: {sheet_name} ===')
    print(f'Dimensions: {ws.dimensions}')
    print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
    merged = list(ws.merged_cells.ranges)
    print(f'Merged cells ({len(merged)}): {merged[:20]}')
    print()
    
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), values_only=False):
        vals = []
        for cell in row:
            v = cell.value
            if v is not None:
                vals.append(f'[{cell.coordinate}]={repr(v)}')
        if vals:
            sep = ' | '
            print(f'Row {row[0].row}: {sep.join(vals)}')
